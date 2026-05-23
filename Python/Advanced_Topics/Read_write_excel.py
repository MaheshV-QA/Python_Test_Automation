"""
-->To read and write Excel files in Python, you can use the pandas and openpyxl libraries.
-->The pandas library provides a DataFrame class that can be used to read and write data to Excel files.
-->The openpyxl library is used to read and write Excel files in the .xlsx format.
-->To install these libraries, you can use the following commands:
-->pip install pandas
-->pip install openpyxl

########################################################
Write Data	=======>df.to_excel("file.xlsx")
Read Data	=======>pd.read_excel("file.xlsx")
Write to Sheet=====>	df.to_excel("file.xlsx", sheet_name="Sheet1")
Read from Sheet=====>	pd.read_excel("file.xlsx", sheet_name="Sheet1")
Append Data	ExcelWriter(mode='a')
Formatting	openpyxl.styles.Font()

"""

import pandas as pd
from openpyxl import Workbook
import os

def create_sheet(file_name):
    wb = Workbook()  # create workbook
    ws = wb.active
    
    wb.save(file_name)
    print("Excel sheet created successfully!")

def write_excel(data, file_name):
    df = pd.DataFrame(data)  # Corrected DataFrame capitalization
    df.to_excel(file_name, index=False)  # index=False prevents writing row indices.
    print(f"Data saved to {file_name}")

def read_excel(file_name):
    df = pd.read_excel(file_name)
    print(df)

if __name__ == "__main__":

    file_name = os.path.join(os.getcwd(), "Python","Advanced_Topics", "sample_excel_data.xlsx")
    print(file_name)

        # Example data to write to Excel
    data = {
        "Name": ["John", "Alice", "Bob"],
        "Age": [28, 24, 30],
        "City": ["New York", "Los Angeles", "Chicago"]
    }
    
    # Create a new Excel sheet
    create_sheet(file_name)
        
    # Write data to Excel
    write_excel(data, file_name)
    
    # Read data from Excel
    read_excel(file_name)